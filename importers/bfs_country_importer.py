#!/usr/bin/env python3
"""Import BFS country codes from official Excel file.

This script reads the official BFS country code Excel file and generates
Python modules with all country data including multilingual names.

Source: sources/bfs/be-b-00.04-sg-01.xlsx (official BFS data)
Output: generated/bfs/country_codes.py

Usage:
    python importers/bfs_country_importer.py
"""

import sys
from pathlib import Path
from typing import Dict, Any, Optional

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openmun-opendata[dev]")
    sys.exit(1)

from importers.base import BaseImporter


class BFSCountryImporter(BaseImporter):
    """Importer for BFS country codes."""

    def import_country_codes(self, force: bool = False) -> bool:
        """Import country codes from Excel and generate Python module.

        Args:
            force: Force regeneration even if output exists

        Returns:
            True if import was successful
        """
        # Paths
        source_file = "be-b-00.04-sg-01.xlsx"
        excel_path = self.sources_dir / "bfs" / source_file
        output_dir = self.generated_dir / "bfs"
        output_file = output_dir / "country_codes.py"

        # Check if regeneration needed
        if not force and output_file.exists():
            print(f"Output already exists: {output_file}")
            print("Use --force to regenerate")
            return True

        # Check input file
        if not excel_path.exists():
            print(f"ERROR: Source file not found: {excel_path}")
            return False

        # Read metadata
        metadata = self.read_metadata("bfs")
        file_metadata = metadata.get(source_file, {})

        print(f"Reading {excel_path}...")
        wb = load_workbook(excel_path, read_only=True)

        # Check for expected sheet
        sheet_name = 'Stat_Geb'
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
            return False

        ws = wb[sheet_name]

        # Parse the data
        countries = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # Skip header
                continue

            bfs_code = row[0]  # Column A: BFS code
            iso2 = row[2]      # Column C: ISO2
            iso3 = row[3]      # Column D: ISO3
            name_de = row[4]   # Column E: German short form
            name_fr = row[5]   # Column F: French short form
            name_it = row[6]   # Column G: Italian short form
            name_en = row[7]   # Column H: English short form
            valid = row[22]    # Column W: Entry valid (J/N)

            # Only include valid entries with both BFS code and ISO2 code
            if valid == 'J' and bfs_code and iso2:
                countries[iso2] = {
                    'bfs_code': str(bfs_code),
                    'iso3': iso3 if iso3 else '',
                    'names': {
                        'de': name_de if name_de else '',
                        'fr': name_fr if name_fr else '',
                        'it': name_it if name_it else '',
                        'en': name_en if name_en else ''
                    }
                }

        print(f"Parsed {len(countries)} valid countries")

        # Ensure output directories exist
        self.ensure_init_files(output_dir)

        # Generate Python module
        print(f"Writing {output_file}...")
        with open(output_file, 'w', encoding='utf-8') as f:
            # Header
            f.write(self.generate_python_header(
                source_file=source_file,
                description=file_metadata.get('description', 'BFS country codes')
            ))

            # Country codes dictionary
            f.write('# ISO 3166-1 alpha-2 codes mapped to BFS codes and names\n')
            f.write('COUNTRY_CODES = {\n')

            for iso2 in sorted(countries.keys()):
                data = countries[iso2]
                f.write(f"    '{iso2}': {{\n")
                f.write(f"        'bfs_code': '{data['bfs_code']}',\n")
                f.write(f"        'iso3': '{data['iso3']}',\n")
                f.write(f"        'names': {{\n")
                for lang, name in data['names'].items():
                    # Escape quotes in names
                    name_escaped = name.replace("'", "\\'")
                    f.write(f"            '{lang}': '{name_escaped}',\n")
                f.write(f"        }}\n")
                f.write(f"    }},\n")

            f.write('}\n\n')

            # Add helper functions
            f.write(self._generate_helper_functions())

        # Write version file
        self.write_version_file(output_dir, source_file, file_metadata)

        print(f"Successfully generated {len(countries)} country codes")
        return True

    def _generate_helper_functions(self) -> str:
        """Generate helper functions for the module."""
        return '''# Type hints
from typing import Optional, Dict, Any

# Helper functions
def get_bfs_country_code(iso_code: str) -> Optional[str]:
    """Get BFS country code from ISO 2-letter code.

    Args:
        iso_code: ISO 3166-1 alpha-2 country code (e.g., 'CH', 'DE')

    Returns:
        BFS country code or None if not found
    """
    country = COUNTRY_CODES.get(iso_code.upper())
    return country['bfs_code'] if country else None


def get_country_name(iso_code: str, language: str = 'de') -> Optional[str]:
    """Get country name in specified language.

    Args:
        iso_code: ISO 3166-1 alpha-2 country code
        language: Language code ('de', 'fr', 'it', 'en')

    Returns:
        Country name or None if not found
    """
    country = COUNTRY_CODES.get(iso_code.upper())
    if country and language in country['names']:
        return country['names'][language]
    return None


def get_country_by_bfs_code(bfs_code: str) -> Optional[Dict[str, Any]]:
    """Get country data by BFS code.

    Args:
        bfs_code: BFS country code (e.g., '8207' for Switzerland)

    Returns:
        Country data dict or None if not found
    """
    bfs_str = str(bfs_code)
    for iso_code, data in COUNTRY_CODES.items():
        if data['bfs_code'] == bfs_str:
            return {'iso2': iso_code, **data}
    return None
'''


def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(description='Import BFS country codes')
    parser.add_argument('--force', action='store_true',
                        help='Force regeneration even if output exists')
    args = parser.parse_args()

    # Run import
    importer = BFSCountryImporter()
    success = importer.import_country_codes(force=args.force)

    if success:
        print("\nDone! You can now import from generated.bfs.country_codes")
    else:
        sys.exit(1)


if __name__ == "__main__":
    main()